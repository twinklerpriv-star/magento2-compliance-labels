import openpyxl
import sys
import os

sys.stdout.reconfigure(encoding='utf-8')

def verify_file(file_path):
    if not os.path.exists(file_path):
        print(f"Error: {file_path} does not exist!")
        return False
        
    print(f"\nVerifying: {file_path}")
    wb = openpyxl.load_workbook(file_path)
    print(f"Sheet names: {wb.sheetnames}")
    ws = wb.active
    
    # Check grid lines
    show_grid = ws.views.sheetView[0].showGridLines
    print(f"Show Grid Lines: {show_grid}")
    if not show_grid:
        print("Warning: Grid lines are not explicitly enabled!")
        
    # Check title
    title = ws.cell(row=1, column=1).value
    print(f"Title: {title}")
    
    # Check headers
    header_row = 4
    headers = []
    col = 1
    while True:
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            break
        headers.append(val)
        
        # Check header styling
        cell = ws.cell(row=header_row, column=col)
        fill = cell.fill
        font = cell.font
        print(f"  Col {col} '{val}': Font Color={font.color.value if font.color else 'None'}, Fill Color={fill.start_color.value if fill.start_color else 'None'}")
        col += 1
        
    print(f"Headers count: {len(headers)}")
    
    # Check input rows styling
    input_row = 5
    sample_cell = ws.cell(row=input_row, column=1)
    print(f"Input row sample fill color: {sample_cell.fill.start_color.value if sample_cell.fill else 'None'}")
    
    return True

def verify_supplier_files(root_dir):
    print(f"\nScanning: {root_dir}")
    if not os.path.exists(root_dir):
        print("Error: root directory does not exist.")
        return
        
    supplier_dirs = [d for d in os.listdir(root_dir) if os.path.isdir(os.path.join(root_dir, d))]
    print(f"Found {len(supplier_dirs)} supplier directories.")
    
    total_de = 0
    total_en = 0
    errors = []
    
    de_headers = [
        "Lie ID",
        "Lieferant Zuname",
        "ArtikelNr",
        "ArtBez1",
        "Marke / Brand (Kurzform für Label - z. B. Miele)",
        "Modellbezeichnung (für Label)",
        "Herstellergarantie (in Jahren)",
        "Link zu Garantiebedingungen (URL beginnend mit https://)",
        "Kostenlos & für gesamte Ware? (Ja/Nein)"
    ]
    
    en_headers = [
        "Supplier ID",
        "Supplier Name",
        "SKU / Article Number",
        "Product Description",
        "Brand (short form for label - e.g. Miele)",
        "Model Identifier (for Label)",
        "Manufacturer Warranty (in years)",
        "Link to Warranty Conditions (URL starting with https://)",
        "Warranty is free of charge & covers entire product? (Yes/No)"
    ]
    
    for s_dir in supplier_dirs:
        dir_path = os.path.join(root_dir, s_dir)
        files = [f for f in os.listdir(dir_path) if f.endswith(".xlsx")]
        
        # Check files in folder
        de_file = f"{s_dir}_Garantieabfrage.xlsx"
        en_file = f"{s_dir}_Warranty_Inquiry.xlsx"
        
        if de_file not in files:
            errors.append(f"Missing German file in {s_dir}")
        if en_file not in files:
            errors.append(f"Missing English file in {s_dir}")
            
        for f in files:
            file_path = os.path.join(dir_path, f)
            is_de = f == de_file
            is_en = f == en_file
            
            if not is_de and not is_en:
                errors.append(f"Unexpected file in {s_dir}: {f}")
                continue
                
            try:
                wb = openpyxl.load_workbook(file_path, read_only=False)
                
                # Check sheet name
                expected_sheet = "Garantiedaten_Erfassung" if is_de else "Warranty_Data_Collection"
                if expected_sheet not in wb.sheetnames:
                    errors.append(f"{f}: Sheet {expected_sheet} not found.")
                    wb.close()
                    continue
                    
                ws = wb[expected_sheet]
                
                # Check gridlines
                show_grid = ws.views.sheetView[0].showGridLines
                if not show_grid:
                    errors.append(f"{f}: Gridlines not enabled.")
                    
                # Check headers (9 columns)
                headers = [ws.cell(row=1, column=c).value for c in range(1, 10)]
                expected_headers = de_headers if is_de else en_headers
                if headers != expected_headers:
                    errors.append(f"{f}: Headers mismatch. Found: {headers}")
                    
                # Check table
                if len(ws.tables) != 1:
                    errors.append(f"{f}: Expected 1 table, found {len(ws.tables)}.")
                else:
                    tab_name = list(ws.tables.keys())[0]
                    tab = ws.tables[tab_name]
                    expected_tab_prefix = "Tabelle_" if is_de else "Tabelle_EN_"
                    if not tab_name.startswith(expected_tab_prefix):
                        errors.append(f"{f}: Table name {tab_name} has invalid prefix.")
                    max_row = ws.max_row
                    if tab.ref != f"A1:I{max_row}":
                        errors.append(f"{f}: Table ref is {tab.ref}, expected A1:I{max_row}.")
                        
                # Check data validation
                dvs = ws.data_validations.dataValidation
                if len(dvs) != 5:
                    errors.append(f"{f}: Expected 5 data validation rules, found {len(dvs)}.")
                else:
                    # Check each column E, F, G, H, I
                    cols_to_check = {
                        "E": {
                            "title": "Marke / Brand" if is_de else "Brand Name",
                            "type": None
                        },
                        "F": {
                            "title": "Modellbezeichnung" if is_de else "Model Identifier",
                            "type": None
                        },
                        "G": {
                            "title": "Herstellergarantie" if is_de else "Manufacturer Warranty",
                            "type": "whole"
                        },
                        "H": {
                            "title": "Link zu Garantiebedingungen" if is_de else "Link to Warranty Conditions",
                            "type": None
                        },
                        "I": {
                            "title": "Garantie-Bestätigung" if is_de else "Warranty Confirmation",
                            "type": "list"
                        }
                    }
                    
                    for col_let, spec in cols_to_check.items():
                        expected_sqref = f"{col_let}2:{col_let}{ws.max_row}"
                        # Find validation rule covering this column
                        found_dv = None
                        for dv in dvs:
                            actual_sqref = str(dv.sqref)
                            if expected_sqref in actual_sqref or (ws.max_row == 2 and actual_sqref == f"{col_let}2"):
                                found_dv = dv
                                break
                        
                        if not found_dv:
                            errors.append(f"{f}: Validation rule for column {col_let} missing.")
                        else:
                            if found_dv.promptTitle != spec["title"]:
                                errors.append(f"{f}: Column {col_let} promptTitle mismatch. Found={found_dv.promptTitle}, Expected={spec['title']}")
                            if spec["type"] and found_dv.type != spec["type"]:
                                errors.append(f"{f}: Column {col_let} validation type mismatch. Found={found_dv.type}, Expected={spec['type']}")
                            if col_let == "G":
                                if found_dv.operator != "greaterThan" or found_dv.formula1 != "2":
                                    errors.append(f"{f}: Column G validation rule mismatch. Op={found_dv.operator}, Formula={found_dv.formula1}")
                            elif col_let == "I":
                                expected_formula = '"Ja,Nein"' if is_de else '"Yes,No"'
                                if found_dv.formula1 != expected_formula:
                                    errors.append(f"{f}: Column I formula mismatch. Found={found_dv.formula1}, Expected={expected_formula}")
                            
                wb.close()
                if is_de:
                    total_de += 1
                else:
                    total_en += 1
            except Exception as e:
                errors.append(f"Error reading {f}: {e}")
                
    print(f"\nVerification Results:")
    print(f"  German files verified: {total_de}")
    print(f"  English files verified: {total_en}")
    print(f"  Total errors: {len(errors)}")
    if errors:
        print("\nErrors details:")
        for err in errors[:20]:
            print(f"  - {err}")
        if len(errors) > 20:
            print(f"  - ... and {len(errors) - 20} more errors.")
    else:
        print("\nAll files verified successfully! No errors found.")

def main():
    target_dir = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Lieferanten"
    verify_supplier_files(target_dir)

if __name__ == "__main__":
    main()
