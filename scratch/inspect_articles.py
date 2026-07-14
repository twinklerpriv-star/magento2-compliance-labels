import openpyxl
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

def inspect_excel(file_path, sheet_name):
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' does not exist.")
        return
        
    print(f"Loading Excel file: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=False)
    
    print(f"Available sheet names: {wb.sheetnames}")
    if sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{sheet_name}' not found.")
        return
        
    ws = wb[sheet_name]
    
    # Read headers (first row)
    headers = []
    first_rows = []
    
    # We read the first few rows to understand the structure
    row_count = 0
    for row in ws.iter_rows(values_only=True):
        row_count += 1
        if row_count == 1:
            headers = list(row)
        elif row_count <= 6:
            first_rows.append(list(row))
        # Keep counting to find total row count
        
    print(f"Total rows counted: {row_count}")
    print("\nHeaders:")
    for idx, header in enumerate(headers, 1):
        print(f"  Col {idx}: {header}")
        
    print("\nSample Data (first 5 rows of data):")
    for idx, row in enumerate(first_rows, 1):
        print(f"  Row {idx}: {row}")

def main():
    file_path = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Meilenstein 1\alle gelisteten Artikel - Stand 24.06.2026.xlsx"
    sheet_name = "Garantiedaten_Erfassung"
    inspect_excel(file_path, sheet_name)

if __name__ == "__main__":
    main()
