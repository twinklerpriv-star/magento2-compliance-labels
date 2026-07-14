import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import shutil
import sys

sys.stdout.reconfigure(encoding='utf-8')

def sanitize_filename(name):
    sanitized = "".join(c if c.isalnum() or c in " .-_" else " " for c in name)
    sanitized = " ".join(sanitized.split())
    sanitized = sanitized.rstrip(". ")
    return sanitized

def sanitize_table_name(name):
    sanitized = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if sanitized and sanitized[0].isdigit():
        sanitized = "_" + sanitized
    return sanitized[:50]

def copy_cell_format(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = src_cell.font
        dest_cell.border = src_cell.border
        dest_cell.fill = src_cell.fill
        dest_cell.number_format = src_cell.number_format
        dest_cell.alignment = src_cell.alignment

def split_by_supplier():
    source_file = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Meilenstein 1\alle gelisteten Artikel - Stand 24.06.2026.xlsx"
    template_de = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Meilenstein 1\Lieferantenabfrage\Supplier_Warranty_Inquiry_Template_DE.xlsx"
    template_en = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Meilenstein 1\Lieferantenabfrage\Supplier_Warranty_Inquiry_Template_EN.xlsx"
    target_root = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Meilenstein 2\Lieferanten"
    
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' does not exist.")
        return
        
    print(f"Loading source data from: {source_file}")
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source["Garantiedaten_Erfassung"]
    
    # Read all rows and group by supplier name
    rows = list(ws_source.iter_rows(values_only=True))
    data_rows = rows[1:]
    
    print(f"Read {len(data_rows)} data rows.")
    
    supplier_data = {}
    for row in data_rows:
        supplier_name = row[1]  # Column 2: Lieferant Zuname
        if not supplier_name:
            continue
        supplier_name_str = str(supplier_name).strip()
        if supplier_name_str not in supplier_data:
            supplier_data[supplier_name_str] = []
        supplier_data[supplier_name_str].append(row)
        
    print(f"Grouped into {len(supplier_data)} suppliers.")
    wb_source.close()
    
    # Clean up old subdirectories under Lieferanten to avoid leftover old files
    if os.path.exists(target_root):
        print("Cleaning up old directories in Lieferanten folder...")
        for name in os.listdir(target_root):
            path = os.path.join(target_root, name)
            if os.path.isdir(path):
                try:
                    shutil.rmtree(path)
                except Exception as e:
                    print(f"Warning: Could not remove old directory {path}: {e}")
                    
    # Generate files for each supplier
    os.makedirs(target_root, exist_ok=True)
    files_created = 0
    
    for supplier_name, items in supplier_data.items():
        sanitized_supplier = sanitize_filename(supplier_name)
        if not sanitized_supplier:
            sanitized_supplier = "Unbekannter_Lieferant"
            
        supplier_folder = os.path.join(target_root, sanitized_supplier)
        os.makedirs(supplier_folder, exist_ok=True)
        
        # -------------------------------------------------------------
        # 1. GERMAN VERSION
        # -------------------------------------------------------------
        de_file_name = f"{sanitized_supplier}_Garantieabfrage.xlsx"
        de_dest_path = os.path.join(supplier_folder, de_file_name)
        shutil.copy(template_de, de_dest_path)
        
        wb_de = openpyxl.load_workbook(de_dest_path)
        ws_de = wb_de["Garantiedaten_Erfassung"]
        
        # Keep row 2 format as template format
        format_cells = [ws_de.cell(row=2, column=col_idx) for col_idx in range(1, 8)]
        
        # Write rows
        for idx, item in enumerate(items):
            row_num = 2 + idx
            
            # Extract EAN/SKU, etc.
            sku = item[2]  # Column 3 of input is SKU
            warranty = item[4]  # Column 5 is warranty
            link = item[5]  # Column 6 is link
            
            # Validate link - only write if it starts with http
            link_str = str(link).strip() if link else ""
            if not link_str.lower().startswith("http"):
                link_str = ""
                
            row_data = [
                None,       # A: EAN (empty)
                sku,        # B: Supplier SKU
                None,       # C: Brand (empty)
                None,       # D: Model (empty)
                warranty,   # E: Warranty
                link_str,   # F: Link
                None        # G: Confirmation (empty)
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_de.cell(row=row_num, column=col_idx, value=val)
                # Apply format from template cell in row 2
                copy_cell_format(format_cells[col_idx-1], cell)
                
        max_row = 1 + len(items)
        if max_row < 2:
            max_row = 2
            
        # Re-create Table
        if "Tabelle1" in ws_de.tables:
            del ws_de.tables["Tabelle1"]
            
        table_name_de = sanitize_table_name(f"Tabelle_{sanitized_supplier}")
        tab_de = Table(displayName=table_name_de, ref=f"A1:G{max_row}")
        style_de = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab_de.tableStyleInfo = style_de
        ws_de.add_table(tab_de)
        
        # Apply German validations
        ws_de.data_validations.dataValidation.clear()
        
        dv_brand = DataValidation(allow_blank=True)
        dv_brand.promptTitle = "Marke / Brand"
        dv_brand.prompt = "Tragen Sie bitte den kurzen Markennamen ein, der auf dem Label stehen soll (z. B. \"Miele\" statt \"Miele GmbH\")."
        ws_de.add_data_validation(dv_brand)
        dv_brand.add(f"C2:C{max_row}")
        
        dv_model = DataValidation(allow_blank=True)
        dv_model.promptTitle = "Modellbezeichnung"
        dv_model.prompt = "Tragen Sie bitte die exakte Modellkennung des Herstellers ein (z. B. \"WXD160\")."
        ws_de.add_data_validation(dv_model)
        dv_model.add(f"D2:D{max_row}")
        
        dv_years = DataValidation(type="whole", operator="greaterThan", formula1=2, allow_blank=True)
        dv_years.promptTitle = "Herstellergarantie"
        dv_years.prompt = "Tragen Sie hier bitte ausschließlich die nackte Zahl ein (z. B. 3, 5, 10). Schreiben Sie keine Einheiten wie \"Jahre\" oder \"J\" dazu."
        dv_years.errorTitle = "Ungültiger Wert"
        dv_years.error = "Geben Sie bitte eine ganze Zahl größer als 2 ein (z.B. 3, 5, 10)."
        ws_de.add_data_validation(dv_years)
        dv_years.add(f"E2:E{max_row}")
        
        dv_link = DataValidation(allow_blank=True)
        dv_link.promptTitle = "Link zu Garantiebedingungen"
        dv_link.prompt = "Bitte tragen Sie hier den direkten Link (URL) zu den Garantiebedingungen auf Ihrer Website ein (z. B. https://www.hersteller.at/garantie). Schreiben Sie keinen Freitext wie \"liegt bei\"."
        ws_de.add_data_validation(dv_link)
        dv_link.add(f"F2:F{max_row}")
        
        dv_confirm = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=True)
        dv_confirm.promptTitle = "Garantie-Bestätigung"
        dv_confirm.prompt = "Bitte wählen Sie ausschließlich \"Ja\" oder \"Nein\" aus dem Dropdown-Menü aus."
        dv_confirm.errorTitle = "Ungültige Auswahl"
        dv_confirm.error = "Ungültige Auswahl. Bitte wählen Sie Ja oder Nein aus dem Dropdown-Menü."
        ws_de.add_data_validation(dv_confirm)
        dv_confirm.add(f"G2:G{max_row}")
        
        # Set widths
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            if col_idx in [1, 2]:
                ws_de.column_dimensions[col_letter].width = 25
            else:
                ws_de.column_dimensions[col_letter].width = 35
                
        wb_de.save(de_dest_path)
        wb_de.close()
        
        # -------------------------------------------------------------
        # 2. ENGLISH VERSION
        # -------------------------------------------------------------
        en_file_name = f"{sanitized_supplier}_Warranty_Inquiry.xlsx"
        en_dest_path = os.path.join(supplier_folder, en_file_name)
        shutil.copy(template_en, en_dest_path)
        
        wb_en = openpyxl.load_workbook(en_dest_path)
        ws_en = wb_en["Warranty_Data_Entry"]
        
        # Keep row 2 format as template format
        format_cells_en = [ws_en.cell(row=2, column=col_idx) for col_idx in range(1, 8)]
        
        # Write rows
        for idx, item in enumerate(items):
            row_num = 2 + idx
            sku = item[2]
            warranty = item[4]
            link = item[5]
            
            link_str = str(link).strip() if link else ""
            if not link_str.lower().startswith("http"):
                link_str = ""
                
            row_data = [
                None,       # A: EAN
                sku,        # B: SKU
                None,       # C: Brand
                None,       # D: Model
                warranty,   # E: Warranty
                link_str,   # F: Link
                None        # G: Confirmation
            ]
            
            for col_idx, val in enumerate(row_data, 1):
                cell = ws_en.cell(row=row_num, column=col_idx, value=val)
                copy_cell_format(format_cells_en[col_idx-1], cell)
                
        # Re-create Table
        if "Table1" in ws_en.tables:
            del ws_en.tables["Table1"]
            
        table_name_en = sanitize_table_name(f"Table_{sanitized_supplier}")
        tab_en = Table(displayName=table_name_en, ref=f"A1:G{max_row}")
        style_en = TableStyleInfo(
            name="TableStyleMedium7",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        tab_en.tableStyleInfo = style_en
        ws_en.add_table(tab_en)
        
        # Apply English validations
        ws_en.data_validations.dataValidation.clear()
        
        dv_brand_en = DataValidation(allow_blank=True)
        dv_brand_en.promptTitle = "Brand"
        dv_brand_en.prompt = "Please enter the short brand name to be shown on the label (e.g. \"Miele\" instead of \"Miele GmbH\")."
        ws_en.add_data_validation(dv_brand_en)
        dv_brand_en.add(f"C2:C{max_row}")
        
        dv_model_en = DataValidation(allow_blank=True)
        dv_model_en.promptTitle = "Model Identifier"
        dv_model_en.prompt = "Please enter the exact manufacturer model code (e.g. \"WXD160\")."
        ws_en.add_data_validation(dv_model_en)
        dv_model_en.add(f"D2:D{max_row}")
        
        dv_years_en = DataValidation(type="whole", operator="greaterThan", formula1=2, allow_blank=True)
        dv_years_en.promptTitle = "Manufacturer Warranty"
        dv_years_en.prompt = "Please enter the warranty period in years as a whole number only (e.g., 3, 5, 10). Do not write \"years\" or \"y\"."
        dv_years_en.errorTitle = "Invalid Value"
        dv_years_en.error = "Please enter a whole number greater than 2 (e.g. 3, 5, 10)."
        ws_en.add_data_validation(dv_years_en)
        dv_years_en.add(f"E2:E{max_row}")
        
        dv_link_en = DataValidation(allow_blank=True)
        dv_link_en.promptTitle = "Link to Warranty Conditions"
        dv_link_en.prompt = "Please enter the direct URL link to the warranty conditions on your website (e.g., https://www.manufacturer.com/warranty). Do not write free text like \"in the box\"."
        ws_en.add_data_validation(dv_link_en)
        dv_link_en.add(f"F2:F{max_row}")
        
        dv_confirm_en = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv_confirm_en.promptTitle = "Warranty Confirmation"
        dv_confirm_en.prompt = "Please select \"Yes\" or \"No\" from the dropdown menu."
        dv_confirm_en.errorTitle = "Invalid Selection"
        dv_confirm_en.error = "Invalid selection. Please select Yes or No from the dropdown menu."
        ws_en.add_data_validation(dv_confirm_en)
        dv_confirm_en.add(f"G2:G{max_row}")
        
        # Set widths
        for col_idx in range(1, 8):
            col_letter = get_column_letter(col_idx)
            if col_idx in [1, 2]:
                ws_en.column_dimensions[col_letter].width = 25
            else:
                ws_en.column_dimensions[col_letter].width = 35
                
        wb_en.save(en_dest_path)
        wb_en.close()
        
        files_created += 1
        if files_created % 10 == 0 or files_created == len(supplier_data):
            print(f"Progress: Generated {files_created}/{len(supplier_data)} supplier packages (DE & EN)...")
            
    print(f"\nSuccessfully generated {files_created} supplier packages (DE & EN) in: {target_root}")

if __name__ == "__main__":
    split_by_supplier()
