import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is not installed. Please install it using 'pip install openpyxl'.")
    sys.exit(1)

def create_styled_sheet(file_path, columns, sheet_name="Garantiedaten", title_text="Garantiedaten Erhebung"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styles
    font_title = Font(name="Segoe UI", size=14, bold=True, color="1A365D")
    font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    font_body = Font(name="Segoe UI", size=11, color="000000")
    font_small = Font(name="Segoe UI", size=9, italic=True, color="555555")
    
    fill_header = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Steel Blue
    fill_input = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Pastel Yellow
    
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    
    # 1. Add Title / Header Info
    ws.cell(row=1, column=1, value=title_text).font = font_title
    ws.row_dimensions[1].height = 25
    
    # Info note depending on language
    info_text = (
        "Bitte tragen Sie hier alle Artikel ein, die eine kostenlose gewerbliche Herstellergarantie von mehr als 2 Jahren aufweisen."
        if "de" in file_path.lower() else
        "Please enter all items that feature a free manufacturer commercial warranty of more than 2 years."
    )
    ws.cell(row=2, column=1, value=info_text).font = font_small
    ws.row_dimensions[2].height = 20
    
    # 2. Add Column Headers
    header_row = 4
    ws.row_dimensions[header_row].height = 28
    
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # 3. Add Empty Styled Input Rows (e.g. 20 rows)
    start_row = 5
    max_row = start_row + 20
    for r in range(start_row, max_row):
        ws.row_dimensions[r].height = 22
        for col_idx in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=col_idx)
            cell.font = font_body
            cell.fill = fill_input
            cell.border = thin_border
            # Alignment matching supplier files
            if col_idx in [1, 2, 3, 7, 9]:
                cell.alignment = Alignment(horizontal="left" if col_idx not in [7, 9] else "center", vertical="center")
            else:
                cell.alignment = Alignment(vertical="center")
                
    # 4. Data Validations
    from openpyxl.worksheet.datavalidation import DataValidation
    is_de = "de" in file_path.lower()
    
    # Validation E: Column E (Brand) - Info Only
    dv_brand = DataValidation(allow_blank=True)
    if is_de:
        dv_brand.prompt = 'Tragen Sie bitte den kurzen Markennamen ein, der auf dem Label stehen soll (z. B. "Miele" statt "Miele GmbH").'
        dv_brand.promptTitle = 'Marke / Brand'
    else:
        dv_brand.prompt = 'Please enter the short brand name to be shown on the label (e.g. "Miele" instead of "Miele GmbH").'
        dv_brand.promptTitle = 'Brand Name'
    ws.add_data_validation(dv_brand)
    dv_brand.add(f"E{start_row}:E{max_row - 1}")
    
    # Validation F: Column F (Model) - Info Only
    dv_model = DataValidation(allow_blank=True)
    if is_de:
        dv_model.prompt = 'Tragen Sie bitte die exakte Modellkennung des Herstellers ein (z. B. "WXD160").'
        dv_model.promptTitle = 'Modellbezeichnung'
    else:
        dv_model.prompt = 'Please enter the exact manufacturer model code (e.g. "WXD160").'
        dv_model.promptTitle = 'Model Identifier'
    ws.add_data_validation(dv_model)
    dv_model.add(f"F{start_row}:F{max_row - 1}")
    
    # Validation G: Column G (Warranty in years) - Ganzzahl > 2
    dv_years = DataValidation(type="whole", operator="greaterThan", formula1="2", allow_blank=True)
    if is_de:
        dv_years.error = 'Geben Sie bitte eine ganze Zahl größer als 2 ein (z.B. 3, 5, 10).'
        dv_years.errorTitle = 'Ungültiger Wert'
        dv_years.prompt = 'Tragen Sie hier bitte ausschließlich die nackte Zahl ein (z. B. 3, 5, 10). Schreiben Sie keine Einheiten wie "Jahre" oder "J" dazu.'
        dv_years.promptTitle = 'Herstellergarantie'
    else:
        dv_years.error = 'Please enter a whole number greater than 2 (e.g. 3, 5, 10).'
        dv_years.errorTitle = 'Invalid Value'
        dv_years.prompt = 'Please enter the warranty period in years as a whole number only (e.g., 3, 5, 10). Do not write "years" or "y".'
        dv_years.promptTitle = 'Manufacturer Warranty'
    ws.add_data_validation(dv_years)
    dv_years.add(f"G{start_row}:G{max_row - 1}")
    
    # Validation H: Column H (URL Link) - Info Only
    dv_link = DataValidation(allow_blank=True)
    if is_de:
        dv_link.prompt = 'Bitte tragen Sie hier den direkten Link (URL) zu den Garantiebedingungen auf Ihrer Website ein (z. B. https://www.hersteller.at/garantie). Schreiben Sie keinen Freitext wie "liegt bei".'
        dv_link.promptTitle = 'Link zu Garantiebedingungen'
    else:
        dv_link.prompt = 'Please enter the direct URL link to the warranty conditions on your website (e.g., https://www.manufacturer.com/warranty). Do not write free text like "in the box".'
        dv_link.promptTitle = 'Link to Warranty Conditions'
    ws.add_data_validation(dv_link)
    dv_link.add(f"H{start_row}:H{max_row - 1}")
    
    # Validation I: Column I (Confirmation) - Dropdown List Ja/Nein or Yes/No
    if is_de:
        dv_list = DataValidation(type="list", formula1='"Ja,Nein"', allow_blank=True)
        dv_list.error = 'Ungültige Auswahl. Bitte wählen Sie Ja oder Nein aus dem Dropdown-Menü.'
        dv_list.errorTitle = 'Ungültige Auswahl'
        dv_list.prompt = 'Bitte wählen Sie ausschließlich "Ja" oder "Nein" aus dem Dropdown-Menü aus.'
        dv_list.promptTitle = 'Garantie-Bestätigung'
    else:
        dv_list = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
        dv_list.error = 'Invalid selection. Please select Yes or No from the dropdown menu.'
        dv_list.errorTitle = 'Invalid Selection'
        dv_list.prompt = 'Please select "Yes" or "No" from the dropdown menu.'
        dv_list.promptTitle = 'Warranty Confirmation'
    ws.add_data_validation(dv_list)
    dv_list.add(f"I{start_row}:I{max_row - 1}")
                
    # 5. Auto-fit column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        # Determine width based on header and sample text
        for cell in col:
            if cell.row == header_row and cell.value:
                max_len = max(max_len, len(str(cell.value)))
                
        # Give columns a safe margin
        ws.column_dimensions[col_letter].width = max(max_len + 5, 20)
        
    wb.save(file_path)
    print(f"Successfully created and styled: {file_path}")

def main():
    target_dir = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL"
    os.makedirs(target_dir, exist_ok=True)
    
    # DE Template columns (9 columns matching the new specification)
    de_columns = [
        "Lie ID",
        "Lieferant Zuname",
        "ArtikelNr",
        "ArtBez1",
        "Marke / Brand (Kurzform für Label - z. B. Miele)",
        "Modellbezeichnung (für Label)",
        "Herstellergarantie (in Jahren)",
        "Link zu Garantiebedingungen (URL beginnend mit https://)",
        "Kostenlos & für gesamte Ware? (Ja/Nein)"
    ]
    de_path = os.path.join(target_dir, "garantieabfrage_lieferanten_de.xlsx")
    create_styled_sheet(de_path, de_columns, "Garantiedaten_Erfassung", "Erfassung der Herstellergarantien (Lieferanten-Abfrage)")
    
    # EN Template columns (9 columns matching the new specification)
    en_columns = [
        "Supplier ID",
        "Supplier Name",
        "SKU / Article Number",
        "Product Description",
        "Brand (short form for label - e.g. Miele)",
        "Model Identifier (for Label)",
        "Manufacturer Warranty (in years)",
        "Link to Warranty Conditions (URL starting with https://)",
        "Warranty is free of charge & covers entire product? (Yes/No)"
    ]
    en_path = os.path.join(target_dir, "warranty_inquiry_suppliers_en.xlsx")
    create_styled_sheet(en_path, en_columns, "Warranty_Data_Collection", "Collection of Manufacturer Warranties (Supplier Inquiry)")

if __name__ == "__main__":
    main()
