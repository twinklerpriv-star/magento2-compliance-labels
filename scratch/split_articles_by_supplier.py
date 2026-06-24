import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
import os
import sys

sys.stdout.reconfigure(encoding='utf-8')

def sanitize_filename(name):
    # Keep alphanumeric, spaces, dots, dashes, and underscores
    sanitized = "".join(c if c.isalnum() or c in " .-_" else " " for c in name)
    # Clean up multiple spaces
    sanitized = " ".join(sanitized.split())
    # Windows folders/files cannot end with a dot or space
    sanitized = sanitized.rstrip(". ")
    return sanitized

def sanitize_table_name(name):
    # Only letters, numbers, and underscores
    sanitized = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    # Ensure it starts with a letter or underscore
    if sanitized and sanitized[0].isdigit():
        sanitized = "_" + sanitized
    # Limit table name length (Excel limit is 255, keep it short and clean)
    return sanitized[:50]

def split_by_supplier():
    source_file = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\alle gelisteten Artikel - Stand 24.06.2026.xlsx"
    template_de = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\scratch\template_de_base.xlsx"
    target_root = r"C:\Users\thomas.winkler\Desktop\Projekte\Google Antigravity\ELEKTROPEPI\GEWAEHRLEISTUNG_GARANTIELABEL\Lieferanten"
    
    if not os.path.exists(source_file):
        print(f"Error: Source file '{source_file}' does not exist.")
        return
        
    print(f"Loading source data from: {source_file}")
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source["Garantiedaten_Erfassung"]
    
    # Read all rows and group by supplier name
    rows = list(ws_source.iter_rows(values_only=True))
    data_rows = rows[1:]
    
    print(f"Read {len(data_rows)} data rows.")
    
    supplier_data = {}
    for row in data_rows:
        supplier_name = row[1] # Column 2: Lieferant Zuname
        if not supplier_name:
            continue
        supplier_name_str = str(supplier_name).strip()
        if supplier_name_str not in supplier_data:
            supplier_data[supplier_name_str] = []
        supplier_data[supplier_name_str].append(row)
        
    print(f"Grouped into {len(supplier_data)} suppliers.")
    wb_source.close()
    
    # Clean up old xlsx files in Lieferanten directory to avoid conflicts
    if os.path.exists(target_root):
        print("Cleaning up old Excel files in Lieferanten directory...")
        for root, dirs, files in os.walk(target_root):
            for file in files:
                if file.endswith(".xlsx"):
                    try:
                        os.remove(os.path.join(root, file))
                    except Exception as e:
                        print(f"Warning: Could not remove old file {file}: {e}")
    
    # Ensure template_de exists
    if not os.path.exists(template_de):
        print("Creating German template from source workbook...")
        wb_temp = openpyxl.load_workbook(source_file)
        ws_temp = wb_temp.active
        ws_temp.delete_rows(2, ws_temp.max_row) # Delete all data rows
        os.makedirs(os.path.dirname(template_de), exist_ok=True)
        wb_temp.save(template_de)
        wb_temp.close()
        
    # Process each supplier
    os.makedirs(target_root, exist_ok=True)
    files_created = 0
    font_header = Font(name="Calibri", size=11, bold=True)
    font_body = Font(name="Calibri", size=11)
    
    headers = [
        "Lie ID",
        "Lieferant Zuname",
        "ArtikelNr",
        "ArtBez1",
        "Marke / Brand (Kurzform für Label - z. B. Miele)",
        "Modellbezeichnung (für Label)",
        "Herstellergarantie (in Jahren)",
        "Link zu Garantiebedingungen (URL beginnend mit https://)",
        "Kostenlos & für gesamte Ware? (Ja/Nein)"
    ]
    
    for supplier_name, items in supplier_data.items():
        sanitized_supplier = sanitize_filename(supplier_name)
        if not sanitized_supplier:
            sanitized_supplier = "Unbekannter_Lieferant"
            
        # Create folder for supplier
        supplier_folder = os.path.join(target_root, sanitized_supplier)
        os.makedirs(supplier_folder, exist_ok=True)
        
        # Load the German base template
        wb_new = openpyxl.load_workbook(template_de)
        ws_new = wb_new.active
        
        # Enable grid lines explicitly
        ws_new.views.sheetView[0].showGridLines = True
        
        # Overwrite header row to ensure correct 9 columns and correct names
        for col_idx, h in enumerate(headers, 1):
            cell = ws_new.cell(row=1, column=col_idx, value=h)
            cell.font = font_header
        
        # Write items (Columns A-D from source, Columns E-I are left blank/None)
        for item in items:
            row_data = [item[0], item[1], item[2], item[3], None, None, None, None, None]
            ws_new.append(row_data)
            
        max_row = len(items) + 1
        
        # Set font and alignments for all data cells
        for r in range(2, max_row + 1):
            for c in range(1, 10):
                cell = ws_new.cell(row=r, column=c)
                cell.font = font_body
                if c in [1, 2, 3, 7, 9]:
                    cell.alignment = Alignment(horizontal="left" if c not in [7, 9] else "center", vertical="center")
                else:
                    cell.alignment = Alignment(vertical="center")
                        
        # Recreate the table from scratch to ensure correct column count, autoFilter range and no corruption
        if ws_new.tables:
            ws_new.tables.clear()
            
            table_name = sanitize_table_name(f"Tabelle_{sanitized_supplier}")
            tab = Table(displayName=table_name, ref=f"A1:I{max_row}")
            style = TableStyleInfo(
                name="TableStyleMedium7",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False
            )
            tab.tableStyleInfo = style
            ws_new.add_table(tab)
        
        # Clear existing and add new validations
        ws_new.data_validations.dataValidation.clear()
        
        # Data Validation E: Column E (Brand) - Info Only
        dv_brand = DataValidation(allow_blank=True)
        dv_brand.prompt = 'Tragen Sie bitte den kurzen Markennamen ein, der auf dem Label stehen soll (z. B. "Miele" statt "Miele GmbH").'
        dv_brand.promptTitle = 'Marke / Brand'
        ws_new.add_data_validation(dv_brand)
        dv_brand.add(f"E2:E{max_row}")
        
        # Data Validation F: Column F (Model) - Info Only
        dv_model = DataValidation(allow_blank=True)
        dv_model.prompt = 'Tragen Sie bitte die exakte Modellkennung des Herstellers ein (z. B. "WXD160").'
        dv_model.promptTitle = 'Modellbezeichnung'
        ws_new.add_data_validation(dv_model)
        dv_model.add(f"F2:F{max_row}")
        
        # Data Validation G: Column G (Warranty in years) - Ganzzahl > 2
        dv_years = DataValidation(
            type="whole",
            operator="greaterThan",
            formula1="2",
            allow_blank=True
        )
        dv_years.error = 'Geben Sie bitte eine ganze Zahl größer als 2 ein (z.B. 3, 5, 10).'
        dv_years.errorTitle = 'Ungültiger Wert'
        dv_years.prompt = 'Tragen Sie hier bitte ausschließlich die nackte Zahl ein (z. B. 3, 5, 10). Schreiben Sie keine Einheiten wie "Jahre" oder "J" dazu.'
        dv_years.promptTitle = 'Herstellergarantie'
        ws_new.add_data_validation(dv_years)
        dv_years.add(f"G2:G{max_row}")
        
        # Data Validation H: Column H (URL Link) - Info Only
        dv_link = DataValidation(allow_blank=True)
        dv_link.prompt = 'Bitte tragen Sie hier den direkten Link (URL) zu den Garantiebedingungen auf Ihrer Website ein (z. B. https://www.hersteller.at/garantie). Schreiben Sie keinen Freitext wie "liegt bei".'
        dv_link.promptTitle = 'Link zu Garantiebedingungen'
        ws_new.add_data_validation(dv_link)
        dv_link.add(f"H2:H{max_row}")
        
        # Data Validation I: Column I (Confirmation) - Dropdown List Ja/Nein
        dv_list = DataValidation(
            type="list",
            formula1='"Ja,Nein"',
            allow_blank=True
        )
        dv_list.error = 'Ungültige Auswahl. Bitte wählen Sie Ja oder Nein aus dem Dropdown-Menü.'
        dv_list.errorTitle = 'Ungültige Auswahl'
        dv_list.prompt = 'Bitte wählen Sie ausschließlich "Ja" oder "Nein" aus dem Dropdown-Menü aus.'
        dv_list.promptTitle = 'Garantie-Bestätigung'
        ws_new.add_data_validation(dv_list)
        dv_list.add(f"I2:I{max_row}")
        
        # Auto-fit columns with safety margin
        for col_idx, col in enumerate(ws_new.columns, 1):
            max_len = 0
            col_letter = get_column_letter(col_idx)
            for cell in col:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            
            # Default widths or auto-fit
            if col_idx in [5, 6, 7, 8]:
                width = max(max_len + 3, 35)
            elif col_idx == 9:
                width = max(max_len + 3, 40)
            else:
                width = max(max_len + 4, 12)
                
            ws_new.column_dimensions[col_letter].width = width
            
        # Save workbook
        file_name = f"{sanitized_supplier}_Garantieabfrage.xlsx"
        file_path = os.path.join(supplier_folder, file_name)
        wb_new.save(file_path)
        wb_new.close()
        files_created += 1
        
        if files_created % 10 == 0 or files_created == len(supplier_data):
            print(f"Progress: Generated {files_created}/{len(supplier_data)} supplier files (DE)...")
            
    print(f"\nSuccessfully generated {files_created} files (DE) in: {target_root}")

if __name__ == "__main__":
    split_by_supplier()
